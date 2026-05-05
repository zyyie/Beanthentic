"""
Import coffee farmer data from Excel file.

This script reads coffee farmer data from an Excel file and exports it
to both JSON and JavaScript formats for use in the application.
"""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None  # type: ignore


def norm_yesno(v) -> str:
    """Normalize yes/no values."""
    if v is None:
        return ""
    s = str(v).strip().lower()
    if s in {"yes", "y", "1", "true"}:
        return "yes"
    if s in {"no", "n", "0", "false"}:
        return "no"
    return s


def fmt_date(v) -> str:
    """Format date value."""
    if v is None or v == "":
        return ""
    if isinstance(v, datetime):
        return f"{v.month}/{v.day}/{v.year}"
    return str(v).strip()


def num(v, default=0.0) -> float:
    """Convert value to float."""
    if v is None or v == "":
        return float(default)
    try:
        return float(v)
    except (ValueError, TypeError):
        try:
            return float(str(v).replace(",", ""))
        except (ValueError, TypeError):
            return float(default)


def intnum(v, default=0) -> int:
    """Convert value to integer."""
    try:
        return int(round(num(v, default)))
    except (ValueError, TypeError):
        return int(default)


def main() -> int:
    """Main function to import coffee database."""
    if load_workbook is None:
        raise ImportError("openpyxl is required to run this script")

    repo_root = Path(__file__).resolve().parents[1]

    xlsx = Path(r"C:\Users\ARLYN\Downloads\coffee-database.xlsx")
    if not xlsx.exists():
        raise FileNotFoundError(f"Missing Excel file: {xlsx}")

    json_path = repo_root / "data" / "farmer-data.json"
    js_path = repo_root / "data" / "farmer-data.js"

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Based on the provided file: headers at row 6; subheaders at 7-8; data starts at 9.
    data_start_row = 9

    # Column mapping by index (1-based) from the Excel layout.
    col = {
        "NO.": 1,
        "NAME OF FARMER": 2,
        "ADDRESS (BARANGAY)": 3,
        "FA OFFICER / MEMBER": 4,
        "BIRTHDAY": 5,
        "RSBSA Registered (Yes/No)": 6,
        # Ownership flags A-E
        "OWNER_OPERATOR": 7,
        "LESSOR": 8,
        "LESSEE": 9,
        "SHAREHOLDER": 10,
        "OTHERS": 11,
        "Total Area Planted (HA.)": 12,
        "LIBERICA BEARING": 13,
        "LIBERICA NON-BEARING": 14,
        "EXCELSA BEARING": 15,
        "EXCELSA NON-BEARING": 16,
        "ROBUSTA BEARING": 17,
        "ROBUSTA NON-BEARING": 18,
        "TOTAL BEARING": 19,
        "TOTAL NON-BEARING": 20,
        "TOTAL TREES": 21,
        "LIBERICA PRODUCTION": 22,
        "EXCELSA PRODUCTION": 23,
        "ROBUSTA PRODUCTION": 24,
        "NCFRS": 25,
        "REMARKS": 26,
    }

    existing: list[dict] = []
    if json_path.exists():
        existing_raw = json.loads(json_path.read_text(encoding="utf-8"))
        if isinstance(existing_raw, list):
            existing = [r for r in existing_raw if isinstance(r, dict)]

    existing_nos = {
        int(r.get("NO.", 0))
        for r in existing
        if isinstance(r.get("NO."), (int, float)) or str(r.get("NO.", "")).strip()
    }

    new_rows: list[dict] = []
    for r in range(data_start_row, ws.max_row + 1):
        no_val = ws.cell(r, col["NO."]).value
        if no_val is None or str(no_val).strip() == "":
            continue
        try:
            no_i = int(float(no_val))
        except (ValueError, TypeError):
            continue

        if no_i in existing_nos:
            continue

        def sval(cname: str) -> str:
            v = ws.cell(r, col[cname]).value
            if v is None:
                return ""
            return v.strip() if isinstance(v, str) else str(v)

        owner_vals = {
            k: ("X" if str(ws.cell(r, col[k]).value or "").strip().upper() == "X" else "")
            for k in ["OWNER_OPERATOR", "LESSOR", "LESSEE", "SHAREHOLDER", "OTHERS"]
        }
        status_ownership = "X" if any(v == "X" for v in owner_vals.values()) else ""

        row = {
            "NO.": no_i,
            "NAME OF FARMER": sval("NAME OF FARMER"),
            "ADDRESS (BARANGAY)": sval("ADDRESS (BARANGAY)"),
            "FA OFFICER / MEMBER": sval("FA OFFICER / MEMBER"),
            "BIRTHDAY": fmt_date(ws.cell(r, col["BIRTHDAY"]).value),
            "RSBSA Registered (Yes/No)": norm_yesno(
                ws.cell(r, col["RSBSA Registered (Yes/No)"]).value
            ),
            "STATUS OF OWNERSHIP": status_ownership,
            "OWNER_OPERATOR": owner_vals["OWNER_OPERATOR"],
            "LESSOR": owner_vals["LESSOR"],
            "LESSEE": owner_vals["LESSEE"],
            "SHAREHOLDER": owner_vals["SHAREHOLDER"],
            "OTHERS": owner_vals["OTHERS"],
            "Total Area Planted (HA.)": num(
                ws.cell(r, col["Total Area Planted (HA.)"]).value, 0.0
            ),
            "LIBERICA BEARING": intnum(ws.cell(r, col["LIBERICA BEARING"]).value, 0),
            "LIBERICA NON-BEARING": intnum(ws.cell(r, col["LIBERICA NON-BEARING"]).value, 0),
            "EXCELSA BEARING": intnum(ws.cell(r, col["EXCELSA BEARING"]).value, 0),
            "EXCELSA NON-BEARING": intnum(ws.cell(r, col["EXCELSA NON-BEARING"]).value, 0),
            "ROBUSTA BEARING": intnum(ws.cell(r, col["ROBUSTA BEARING"]).value, 0),
            "ROBUSTA NON-BEARING": intnum(ws.cell(r, col["ROBUSTA NON-BEARING"]).value, 0),
            "TOTAL BEARING": intnum(ws.cell(r, col["TOTAL BEARING"]).value, 0),
            "TOTAL NON-BEARING": intnum(ws.cell(r, col["TOTAL NON-BEARING"]).value, 0),
            "TOTAL TREES": intnum(ws.cell(r, col["TOTAL TREES"]).value, 0),
            "LIBERICA PRODUCTION": num(ws.cell(r, col["LIBERICA PRODUCTION"]).value, 0.0),
            "EXCELSA PRODUCTION": num(ws.cell(r, col["EXCELSA PRODUCTION"]).value, 0.0),
            "ROBUSTA PRODUCTION": num(ws.cell(r, col["ROBUSTA PRODUCTION"]).value, 0.0),
            "NCFRS": sval("NCFRS"),
            "REMARKS": sval("REMARKS"),
        }

        new_rows.append(row)

    merged = existing + new_rows
    merged.sort(key=lambda d: int(d.get("NO.", 0) or 0))

    json_path.write_text(json.dumps(merged, indent=2, ensure_ascii=False), encoding="utf-8")

    js_content = (
        "// Farmer data from Excel file\n"
        "const farmerData = "
        + json.dumps(merged, indent=2, ensure_ascii=False)
        + ";\n\n"
        "// Export the data for use in the dashboard\n"
        "if (typeof module !== 'undefined' && module.exports) {\n"
        "  module.exports = farmerData;\n"
        "} else {\n"
        "  window.farmerData = farmerData;\n"
        "}\n"
    )
    js_path.write_text(js_content, encoding="utf-8")

    print(f"Imported from: {xlsx}")
    print(f"Existing records: {len(existing)}")
    print(f"New records added: {len(new_rows)}")
    print(f"Total records now: {len(merged)}")
    print(f"Updated: {json_path}")
    print(f"Updated: {js_path}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
